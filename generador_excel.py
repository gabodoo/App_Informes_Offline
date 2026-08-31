import sys
from pathlib import Path
import openpyxl

def get_base_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent

class GeneradorExcel:
    """Clase para la generación de la Notificación Excel a partir de su plantilla."""

    def __init__(self):
        self.ruta_plantilla = None

    def generar(self, contexto: dict, log_callback=None) -> Path:
        def log(msg):
            if log_callback:
                log_callback(msg)

        es_casilla = contexto.get("AUTORIZA_CASILLA", False)
        if es_casilla:
            self.ruta_plantilla = get_base_dir() / "plantillas" / "plantilla_notificacion_casilla.xlsx"
        else:
            self.ruta_plantilla = get_base_dir() / "plantillas" / "plantilla_notificacion.xlsx"

        if not self.ruta_plantilla.exists():
            log(f"No se encontró la plantilla de notificación en {self.ruta_plantilla}")
            raise FileNotFoundError(f"Falta plantilla: {self.ruta_plantilla}")

        log("Cargando plantilla de notificación Excel...")
        wb = openpyxl.load_workbook(self.ruta_plantilla)
        ws = wb.active
        
        sigedo_corto = contexto.get("SIGEDO_CORTO", "")
        dni = contexto.get("DNI_VALIDADO", "")
        n_bec = str(contexto.get("NOMBRES_BECARIO", "")).strip()
        a_bec = str(contexto.get("APELLIDOS_BECARIO", "")).strip()
        nombres = f"{n_bec} {a_bec}".strip()
        if not nombres:
            nombres = str(contexto.get("NOMBRES_Y_APELLIDOS_VALIDADOS", "")).upper()
        correo = contexto.get("CORREO_ELECTRONICO", "")
        informe_generado = f"INFORME Nº {contexto.get('NUMERO_INFORME_GENERAR', '')}-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS"

        if es_casilla:
            ws["B3"].value = dni
            ws["C3"].value = nombres
            ws["D3"].value = contexto.get("EXPEDIENTE_BECARIO", "")
            ws["E3"].value = contexto.get("NUMERO_SIGEDO", "")
            ws["K3"].value = contexto.get("TELEFONO_CONTACTO", "")
            
            import re
            texto_doc = str(ws["H3"].value or "")
            texto_doc = re.sub(r"INFORME N[°º]\s*\d+-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS", informe_generado, texto_doc, flags=re.IGNORECASE)
            ws["H3"].value = texto_doc
            
            nombre_salida = f"{sigedo_corto}_notificacion_casilla.xlsx"
        else:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    texto_original = str(cell.value)
                    
                    # Reemplazo SIGEDO (59583)
                    if "59583" in texto_original or "60692" in texto_original:
                        texto_a_reemplazar = "59583" if "59583" in texto_original else "60692"
                        if isinstance(cell.value, (int, float)):
                            import re
                            solo_nums = re.sub(r"\D", "", str(sigedo_corto))
                            cell.value = int(solo_nums) if solo_nums else sigedo_corto
                        else:
                            cell.value = texto_original.replace(texto_a_reemplazar, str(sigedo_corto))
                        
                    # Reemplazo de Informe
                    if "6539" in str(cell.value) or "6757" in str(cell.value):
                        if isinstance(cell.value, (int, float)):
                            num = contexto.get('NUMERO_INFORME_GENERAR', '')
                            cell.value = int(num) if str(num).isdigit() else num
                        else:
                            pass # Será manejado por las siguientes condiciones si es un string largo
                            
                    texto = str(cell.value)
                    
                    if "Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS" in texto:
                        cell.value = texto.replace("Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS", informe_generado)
                    elif "Informe N° 6757-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS" in texto:
                        cell.value = texto.replace("Informe N° 6757-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS", informe_generado)
                    elif "Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC" in texto:
                        cell.value = texto.replace("Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC", informe_generado)
                    elif "Informe N° 6539" in texto:
                        cell.value = texto.replace("Informe N° 6539", informe_generado)
                    elif "Informe N° 6757" in texto:
                        cell.value = texto.replace("Informe N° 6757", informe_generado)
                    
                    # Reemplazos "debajo" de una celda
                    texto_upper = texto.upper()
                    if "DNI" == texto_upper.strip() or "D.N.I." in texto_upper:
                        celda_abajo = ws.cell(row=cell.row + 1, column=cell.column)
                        celda_abajo.value = dni
                    
                    if "DESTINATARIO" == texto_upper.strip() or "DESTINATARIO" in texto_upper and "NOTIFICACIONES" not in texto_upper:
                        celda_abajo = ws.cell(row=cell.row + 1, column=cell.column)
                        celda_abajo.value = nombres
                        
                    if ("CORREO ELECTRÓNICO" in texto_upper or "CORREO ELECTRONICO" in texto_upper) and "NOTIFICACIONES" not in texto_upper:
                        celda_abajo = ws.cell(row=cell.row + 1, column=cell.column)
                        celda_abajo.value = correo

            nombre_salida = f"{sigedo_corto}_notificacion.xlsx"
        ruta_salida = get_base_dir() / "Informes_Generados" / nombre_salida
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        
        log(f"Guardando Notificación en: {ruta_salida.name}")
        wb.save(ruta_salida)
        return ruta_salida

    def generar_multiple(self, super_contexto: dict, log_callback=None) -> Path:
        def log(msg):
            if log_callback: log_callback(msg)

        self.ruta_plantilla = get_base_dir() / "plantillas" / "plantilla_notificacion_multiple.xlsx"
        if not self.ruta_plantilla.exists():
            log(f"No se encontró la plantilla múltiple en {self.ruta_plantilla}")
            raise FileNotFoundError(f"Falta plantilla: {self.ruta_plantilla}")

        log("Cargando plantilla de notificación Excel múltiple...")
        wb = openpyxl.load_workbook(self.ruta_plantilla)
        ws = wb.active
        
        # Guardamos el formato y valores base de la primera fila de datos (asumimos fila 3)
        fila_base = 3
        valores_base = [ws.cell(row=fila_base, column=c).value for c in range(1, 14)]
        
        # Limpiar filas posteriores si existen en la plantilla
        for r in range(ws.max_row, fila_base, -1):
            ws.delete_rows(r)
            
        num_inf = super_contexto.get("NUMERO_INFORME_GENERAR", "")
        informe_generado = f"INFORME N° {num_inf}-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS"
        sigedo_global = super_contexto.get("NUMERO_SIGEDO_GLOBAL", "")

        for i, becario in enumerate(super_contexto.get("becarios", [])):
            fila_actual = fila_base + i
            
            dni = becario.get("DNI_VALIDADO", "")
            n_bec = str(becario.get("NOMBRES_BECARIO", "")).strip()
            a_bec = str(becario.get("APELLIDOS_BECARIO", "")).strip()
            nombres = f"{n_bec} {a_bec}".strip()
            if not nombres:
                nombres = str(becario.get("NOMBRES_Y_APELLIDOS_VALIDADOS", "")).upper()
            correo = becario.get("CORREO_ELECTRONICO", "")
            tel = becario.get("TELEFONO_CONTACTO", "")
            
            # Col A: N°
            ws.cell(row=fila_actual, column=1, value=i+1)
            # Col B: SIGEDO
            ws.cell(row=fila_actual, column=2, value=sigedo_global)
            # Col C: DNI
            ws.cell(row=fila_actual, column=3, value=dni)
            # Col D: OFICINA REMITENTE
            ws.cell(row=fila_actual, column=4, value=valores_base[3] if len(valores_base) > 3 else "SUBDIRECCIÓN DE SEGUIMIENTO Y SUPERVISIÓN")
            # Col E: DESTINATARIO
            ws.cell(row=fila_actual, column=5, value=nombres)
            # Col F: TIPIFICACIÓN
            ws.cell(row=fila_actual, column=6, value=valores_base[5] if len(valores_base) > 5 else "")
            # Col G: N° DE DOCUMENTO
            ws.cell(row=fila_actual, column=7, value=valores_base[6] if len(valores_base) > 6 else "")
            
            # Col H: DESCRIPCIÓN
            desc_original = str(valores_base[7] if len(valores_base) > 7 else "")
            import re
            desc_mod = re.sub(r"INFORME N[°ºo]\s*\d+-\d+-MINEDU/VMGI-PRONABEC-DIBEC-SUS", informe_generado, desc_original, flags=re.IGNORECASE)
            ws.cell(row=fila_actual, column=8, value=desc_mod)
            
            # Col I: CORREO
            ws.cell(row=fila_actual, column=9, value=correo)
            # Col J: TELEFONO
            ws.cell(row=fila_actual, column=10, value=tel)
            # Col K: ITEM ARCHIVO
            ws.cell(row=fila_actual, column=11, value=i+1)
            # Col L: OBSERVACIONES
            ws.cell(row=fila_actual, column=12, value=valores_base[11] if len(valores_base) > 11 else "")
            # Col M: TIPO DE NOTIFICACIÓN
            ws.cell(row=fila_actual, column=13, value=valores_base[12] if len(valores_base) > 12 else "2. COMUNICACIÓN")

        nombre_salida = f"{sigedo_global}_notificacion_multiple.xlsx"
        ruta_salida = get_base_dir() / "Informes_Generados" / nombre_salida
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        
        log(f"Guardando Notificación Múltiple en: {ruta_salida.name}")
        wb.save(ruta_salida)
        return ruta_salida

