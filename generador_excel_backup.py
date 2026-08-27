import sys
from pathlib import Path
import openpyxl

def get_base_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent

class GeneradorExcel:
    """Clase para la generación de la Notificación Excel a partir de su plantilla."""

    def __init__(self):
        self.ruta_plantilla = get_base_dir() / "plantillas" / "plantilla_notificacion.xlsx"

    def generar(self, contexto: dict, log_callback=None) -> Path:
        def log(msg):
            if log_callback:
                log_callback(msg)

        if not self.ruta_plantilla.exists():
            log(f"No se encontró la plantilla de notificación en {self.ruta_plantilla}")
            raise FileNotFoundError(f"Falta plantilla: {self.ruta_plantilla}")

        log("Cargando plantilla de notificación Excel...")
        wb = openpyxl.load_workbook(self.ruta_plantilla)
        ws = wb.active
        
        sigedo_corto = contexto.get("SIGEDO_CORTO", "")
        dni = contexto.get("DNI_VALIDADO", "")
        n_bec = str(contexto.get("NOMBRES_BECARIO", "")).strip()
        a_bec = str(contexto.get("APELLIDOS_BECARIO", "")).strip()
        nombres = f"{n_bec} {a_bec}".strip()
        if not nombres:
            nombres = str(contexto.get("NOMBRES_Y_APELLIDOS_VALIDADOS", "")).upper()
        correo = contexto.get("CORREO_ELECTRONICO", "")
        informe_generado = f"INFORME Nº {contexto.get('NUMERO_INFORME_GENERAR', '')}-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS"

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                texto_original = str(cell.value)
                
                # Reemplazo SIGEDO (59583)
                if "59583" in texto_original or "60692" in texto_original:
                    texto_a_reemplazar = "59583" if "59583" in texto_original else "60692"
                    if isinstance(cell.value, (int, float)):
                        import re
                        solo_nums = re.sub(r"\D", "", str(sigedo_corto))
                        cell.value = int(solo_nums) if solo_nums else sigedo_corto
                    else:
                        cell.value = texto_original.replace(texto_a_reemplazar, str(sigedo_corto))
                    
                # Reemplazo de Informe
                if "6539" in str(cell.value) or "6757" in str(cell.value):
                    if isinstance(cell.value, (int, float)):
                        num = contexto.get('NUMERO_INFORME_GENERAR', '')
                        cell.value = int(num) if str(num).isdigit() else num
                    else:
                        pass # Será manejado por las siguientes condiciones si es un string largo
                        
                texto = str(cell.value)
                
                if "Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS" in texto:
                    cell.value = texto.replace("Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS", informe_generado)
                elif "Informe N° 6757-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS" in texto:
                    cell.value = texto.replace("Informe N° 6757-2026-MINEDU/VMGI-PRONABEC-DIBEC-SUS", informe_generado)
                elif "Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC" in texto:
                    cell.value = texto.replace("Informe N° 6539-2026-MINEDU/VMGI-PRONABEC-DIBEC", informe_generado)
                elif "Informe N° 6539" in texto:
                    cell.value = texto.replace("Informe N° 6539", informe_generado)
                elif "Informe N° 6757" in texto:
                    cell.value = texto.replace("Informe N° 6757", informe_generado)
                
                # Reemplazos "debajo" de una celda
                texto_upper = texto.upper()
                if "DNI" == texto_upper.strip() or "D.N.I." in texto_upper:
                    celda_abajo = ws.cell(row=cell.row + 1, column=cell.column)
                    celda_abajo.value = dni
                
                if "DESTINATARIO" == texto_upper.strip() or "DESTINATARIO" in texto_upper and "NOTIFICACIONES" not in texto_upper:
                    celda_abajo = ws.cell(row=cell.row + 1, column=cell.column)
                    celda_abajo.value = nombres
                    
                if ("CORREO ELECTRÓNICO" in texto_upper or "CORREO ELECTRONICO" in texto_upper) and "NOTIFICACIONES" not in texto_upper:
                    celda_abajo = ws.cell(row=cell.row + 1, column=cell.column)
                    celda_abajo.value = correo

        nombre_salida = f"{sigedo_corto}_notificacion.xlsx"
        ruta_salida = get_base_dir() / "Informes_Generados" / nombre_salida
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        
        log(f"Guardando Notificación en: {ruta_salida.name}")
        wb.save(ruta_salida)
        return ruta_salida
